

import cv2  # OpenCV for image processing
import imutils  # For image resizing
import pytesseract  # Tesseract OCR for text recognition
import pandas as pd  # For creating and managing data frames
from openpyxl import Workbook  # For Excel enhancements
from openpyxl.styles import Font, Alignment  # For Excel formatting

# Set the path to the Tesseract executable (Update this path based on your installation)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Step 1: Capture video from the webcam
vid = cv2.VideoCapture(0)  # '0' indicates the default camera

while True:
    ret, frame = vid.read()  # Capture frame-by-frame
    if not ret:  # Check if the frame is captured
        print("Failed to grab frame. Check your webcam connection.")
        break
    
    cv2.imshow('Webcam Feed (Press "q" to capture)', frame)  # Display the frame
    
    # Press 'q' to capture the photo and exit the loop
    if cv2.waitKey(1) & 0xFF == ord('q'):
        cv2.imwrite('captured_plate.jpg', frame)  # Save the captured frame
        break

vid.release()  # Release the webcam
cv2.destroyAllWindows()  # Close all OpenCV windows

# Step 2: Read and preprocess the captured image
image = cv2.imread('captured_plate.jpg')
image = imutils.resize(image, width=500)  # Resize the image for better processing

# Convert the image to grayscale
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
gray = cv2.bilateralFilter(gray, 11, 17, 17)  # Reduce noise while preserving edges
edged = cv2.Canny(gray, 170, 200)  # Detect edges using the Canny method

# Step 3: Find and process contours
cnts, _ = cv2.findContours(edged.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:30]  # Sort and keep the largest 30 contours
NumberPlateCnt = None  # Initialize the number plate contour

for i in cnts:
    perimeter = cv2.arcLength(i, True)
    approx = cv2.approxPolyDP(i, 0.02 * perimeter, True)
    
    if len(approx) == 4:  # Detect rectangle (possible number plate)
        NumberPlateCnt = approx
        x, y, w, h = cv2.boundingRect(i)  # Get the bounding box
        cropped_image = image[y:y + h, x:x + w]  # Crop the detected plate region
        cv2.imwrite('number_plate.png', cropped_image)  # Save the cropped image
        break

# Step 4: Perform OCR if a number plate is detected
if NumberPlateCnt is not None:
    # Extract text from the cropped number plate image
    text = pytesseract.image_to_string('number_plate.png', lang='eng')
    cleaned_text = ''.join(e for e in text if e.isalnum())  # Clean the text
    
    # Display detected and cleaned text
    print("Detected Text:", text.strip())
    print("Cleaned Text:", cleaned_text)
    
    # Step 5: Save the results to an Excel spreadsheet
    data = {'Detected Text': [text.strip()], 'Cleaned Text': [cleaned_text]}
    df = pd.DataFrame(data)
    
    # Create a new Excel workbook and write data to it
    excel_file = r'C:\Users\Joseph\Downloads\anpr_output.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ANPR Results')
        
        # Access the workbook and sheet to apply formatting
        workbook = writer.book
        sheet = writer.sheets['ANPR Results']
        
        # Set column width and apply formatting
        for column in sheet.columns:
            max_length = 0
            column_name = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_name].width = adjusted_width
        
        # Apply bold and centered formatting to header
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Center-align the data cells
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
    
    print("Data stored in 'number_plate_data.xlsx'.")
else:
    print("No number plate detected.")
